import openpyxl
from openpyxl.styles import Font
import re

wb = openpyxl.load_workbook("Relatorio_GPON_LUMA01.xlsx")
ws = wb.active

col_power = None
col_rssi = None

for cell in ws[1]:
    if cell.value == "Power Level":
        col_power = cell.column
    elif cell.value == "RSSI":
        col_rssi = cell.column

fonte_vermelha = Font(color="FF0000", bold=True)

for row in range(2, ws.max_row + 1):
    for col in (col_power, col_rssi):
        if col:
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value) != "N/A":
                try:
                    match = re.search(r"[-+]?\d*\.\d+|\d+", str(cell.value))
                    if match:
                        valor = float(match.group())
                        if -31.0 <= valor <= -24.0:
                            cell.font = fonte_vermelha
                except ValueError:
                    pass

wb.save("Relatorio_GPON_LUMA01_Formatado.xlsx")