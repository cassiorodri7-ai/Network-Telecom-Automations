import pandas as pd
import glob
import re
import openpyxl
from openpyxl.styles import Font

# ==========================================
# ETAPA 1: EXTRAÇÃO DE DADOS (PARSING)
# ==========================================
print("🔍 Lendo arquivos de texto da OLT...")
arquivos_txt = glob.glob("*.txt")
todas_as_linhas = []

for arquivo in arquivos_txt:
    # Ignora arquivos de dependência ou logs que não são da OLT
    if arquivo == "requirements.txt" or arquivo.endswith(".py"): 
        continue
        
    # Tratamento do nome do arquivo (Ex: gpon14.txt -> gpon1/4)
    nome_base = arquivo.replace('.txt', '')
    nome_olt = re.sub(r'(gpon\d)(\d+)', r'\1/\2', nome_base, flags=re.IGNORECASE)
        
    with open(arquivo, 'r', encoding='utf-8') as f:
        contrato = sn = power = rssi = None
        
        for linha in f:
            linha = linha.strip()
            
            # Regex para capturar o Contrato e o Serial Number
            match_cabecalho = re.match(r'^\d+-([^\(]+)\s*\(([^)]+)\):', linha)
            
            if match_cabecalho:
                if contrato:
                    todas_as_linhas.append({
                        "Arquivo OLT": nome_olt,
                        "Contrato": contrato,
                        "Serial Number": sn,
                        "Power Level": power,
                        "RSSI": rssi
                    })
                
                contrato = match_cabecalho.group(1).strip()
                sn = match_cabecalho.group(2).strip()
                power = "N/A"
                rssi = "N/A"
                
            # Extrai os níveis de sinal
            elif linha.startswith("Power Level"):
                power = linha.split(":")[1].split("(")[0].strip()
                
            elif linha.startswith("RSSI"):
                rssi = linha.split(":")[1].split("(")[0].strip()
        
        # Salva a última ONU do arquivo
        if contrato:
            todas_as_linhas.append({
                "Arquivo OLT": nome_olt,
                "Contrato": contrato,
                "Serial Number": sn,
                "Power Level": power,
                "RSSI": rssi
            })

if not todas_as_linhas:
    print("❌ Nenhum dado no formato GPON foi encontrado nos arquivos .txt.")
else:
    # ==========================================
    # ETAPA 2: CRIAÇÃO DO BANCO DE DADOS
    # ==========================================
    df = pd.DataFrame(todas_as_linhas)
    nome_arquivo = "Relatorio_GPON_Final.xlsx"
    df.to_excel(nome_arquivo, index=False)
    print(f"✅ Extração concluída! Foram processadas {len(todas_as_linhas)} ONUs.")
    
    # ==========================================
    # ETAPA 3: FORMATAÇÃO VISUAL (HIGHLIGHT)
    # ==========================================
    print("🎨 Aplicando formatação de sinais críticos...")
    
    wb = openpyxl.load_workbook(nome_arquivo)
    ws = wb.active

    col_power = None
    col_rssi = None

    # Localiza as colunas dinamicamente para não quebrar se a ordem mudar
    for cell in ws[1]:
        if cell.value == "Power Level":
            col_power = cell.column
        elif cell.value == "RSSI":
            col_rssi = cell.column

    fonte_vermelha = Font(color="FF0000", bold=True)

    # Varre as linhas aplicando a regra matemática de sinal ruim (-24 a -31)
    for row in range(2, ws.max_row + 1):
        for col in (col_power, col_rssi):
            if col:
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value) != "N/A":
                    try:
                        # Extrai apenas o número do texto (ex: tira o "dBm")
                        match = re.search(r"[-+]?\d*\.\d+|\d+", str(cell.value))
                        if match:
                            valor = float(match.group())
                            if valor <= -24.0:
                                cell.font = fonte_vermelha
                    except ValueError:
                        pass

    # Salva o resultado final em cima do mesmo arquivo para manter a pasta limpa
    wb.save(nome_arquivo)
    print(f"✨ Pronto! O arquivo '{nome_arquivo}' foi gerado e formatado com sucesso.")